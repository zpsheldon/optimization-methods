#!/usr/bin/env python3
"""
Update `review_allocation.xlsx` cell fills based on `review_allocation.csv` mask.

Behavior (preserve-original mode):
- For each cell in the target sheet, if the corresponding CSV value == 1, leave the
  cell's fill unchanged (preserve whatever color was there).
- If the CSV value != 1 (or CSV has no corresponding cell), clear the cell's fill
  (set to no background).

By default the script modifies the input `.xlsx` in-place, but it creates a backup
named `<input>.bak.xlsx` before overwriting. You may pass `--output-xlsx` to save
a separate file instead.

Usage example:
    python apply_review_allocation_cf.py \
      --input-xlsx review_allocation.xlsx \
      --mask-csv review_allocation.csv \
      --output-xlsx review_allocation_updated.xlsx
"""
import argparse
from pathlib import Path
import sys
import shutil
import time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--input-xlsx", default="review_allocation.xlsx",
                   help="Input Excel file to modify")
    p.add_argument("--mask-csv", default="review_allocation.csv",
                   help="CSV file containing mask values (1 to keep fill)")
    p.add_argument("--sheet-index", type=int, default=0,
                   help="Zero-based sheet index to operate on (default: 0)")
    p.add_argument("--mask-col-offset", type=int, default=0,
                   help="Column offset to apply when mapping mask CSV columns to sheet columns (default 0).\nExample: use 1 when CSV referee columns are shifted right by one relative to the sheet.")
    p.add_argument("--output-xlsx", default=None,
                   help="If provided, save to this file instead of overwriting input")
    p.add_argument("--backup", action="store_true",
                   help="When overwriting input, create a .bak.xlsx backup (default true)")
    p.add_argument("--fill-color", default="FFFFFF00",
                   help="ARGB hex color to apply for mask==1 (default: FFFF FFFF?), e.g. FFFFEB9C")
    return p.parse_args()


def main():
    args = parse_args()
    input_xlsx = Path(args.input_xlsx)
    mask_csv = Path(args.mask_csv)

    if not input_xlsx.exists():
        print(f"Error: input xlsx not found: {input_xlsx}")
        sys.exit(2)
    if not mask_csv.exists():
        print(f"Error: mask csv not found: {mask_csv}")
        sys.exit(2)

    try:
        mask_df = pd.read_csv(mask_csv, header=None, dtype=object)
    except Exception as e:
        print(f"Error reading mask CSV: {e}")
        sys.exit(2)

    # Convert to boolean mask where cell equals string '1' or numeric 1
    def is_one(v):
        try:
            if pd.isna(v):
                return False
        except Exception:
            pass
        s = str(v).strip()
        return s == "1"

    # Use applymap to produce a DataFrame of booleans where True means keep fill
    # (cell equals '1'). DataFrame.map is not appropriate here.
    mask_bool = mask_df.applymap(is_one)

    wb = load_workbook(filename=str(input_xlsx))
    sheet_names = wb.sheetnames
    if args.sheet_index < 0 or args.sheet_index >= len(sheet_names):
        print(f"Invalid sheet index {args.sheet_index}; workbook has {len(sheet_names)} sheets")
        sys.exit(2)
    ws = wb[sheet_names[args.sheet_index]]

    max_row = ws.max_row
    max_col = ws.max_column
    csv_rows, csv_cols = mask_bool.shape

    if (csv_rows < max_row) or (csv_cols < max_col):
        print(f"Warning: mask CSV shape ({csv_rows},{csv_cols}) is smaller than sheet ({max_row},{max_col}).")
        print("Only overlapping cells will be processed; non-overlapping cells will have fills cleared.")

    # Remove existing conditional formatting so visual colors from CF don't persist
    try:
        # prefer public API if available
        if hasattr(ws.conditional_formatting, 'clear'):
            ws.conditional_formatting.clear()
        else:
            ws.conditional_formatting._cf_rules.clear()
    except Exception:
        # best-effort: ignore if we can't clear
        pass

    # Build fills
    empty_fill = PatternFill()

    # Color mapping for workbook cell values (case-insensitive)
    color_map = {
        'yes': 'FFC6EFCE',      # light green
        'no': 'FFFFC7CE',       # light red
        'maybe': 'FFFFEB9C',    # light yellow
        'conflict': 'FFD9D9D9', # light grey
    }

    def pick_fill_for_cell_value(v):
        if v is None:
            return empty_fill
        s = str(v).strip().lower()
        if s in color_map:
            return PatternFill(fill_type='solid', fgColor=color_map[s])
        return empty_fill

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            keep = False
            # Map sheet column `c` to mask column index using mask-col-offset.
            mask_col_index = (c - 1) + args.mask_col_offset
            if (r - 1) < csv_rows and (mask_col_index) < csv_cols and mask_col_index >= 0:
                try:
                    keep = bool(mask_bool.iat[r - 1, mask_col_index])
                except Exception:
                    val = str(mask_df.iat[r - 1, mask_col_index]).strip() if (r - 1) < csv_rows and (mask_col_index) < csv_cols and mask_col_index >= 0 else ""
                    keep = (val == "1")
            cell = ws.cell(row=r, column=c)
            if keep:
                # choose fill based on the cell's textual value in the workbook
                cell.fill = pick_fill_for_cell_value(cell.value)
            else:
                cell.fill = empty_fill

    # Save workbook
    if args.output_xlsx:
        outpath = Path(args.output_xlsx)
        wb.save(str(outpath))
        print(f"Saved updated workbook to: {outpath}")
    else:
        # overwrite input; create backup
        backup_path = input_xlsx.with_suffix('.bak.xlsx')
        # If a backup already exists, append timestamp
        if backup_path.exists():
            ts = int(time.time())
            backup_path = input_xlsx.with_name(input_xlsx.stem + f'.bak.{ts}.xlsx')
        shutil.copy2(str(input_xlsx), str(backup_path))
        wb.save(str(input_xlsx))
        print(f"Overwrote {input_xlsx}; backup saved as {backup_path}")


if __name__ == '__main__':
    main()
